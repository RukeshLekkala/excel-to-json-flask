from flask import Flask, request, redirect, url_for, render_template

import os
import json
import glob

from openpyxl import Workbook , load_workbook
from openpyxl.utils import column_index_from_string

app = Flask(__name__)

DIR_JSON = os.path.join("uploadr/static/results/")

# READ ALL ROWS IN WORKSHEET AND TRANSFORM INTO JSON
def all_data_to_json(worksheet, filename, sheetname):
    # with open('{}{}_{}.json'.format(DIR_JSON,filename,sheetname), 'w') as file:
    with open('{}{}_{}.json'.format(DIR_JSON, filename,sheetname), 'w') as file:
        max_row = worksheet.max_row
        max_column = worksheet.max_column

        cabecalho = []
        for col in worksheet.iter_rows(min_row=1, max_col=max_column, max_row=1):
            for cell in col:
                cabecalho.append(cell.value)

        data = []
        for row in worksheet.iter_rows(min_row=2, max_col=max_column, max_row=max_row):
            item = {}
            for cell in row: 
                try:
                    if("Date" in cabecalho[column_index_from_string(cell.column)-1] or "Data" in cabecalho[column_index_from_string(cell.column)-1]):
                        item["Date"] = str(cell.value)
                        continue
                except:
                    x = 0
                try:
                    item[cabecalho[column_index_from_string(cell.column)-1].encode("utf-8")] = cell.value.encode("utf-8")
                except:
                    try:
                        item[cabecalho[column_index_from_string(cell.column)-1].encode("utf-8")] = cell.value
                    except:
                        try:
                            item[cabecalho[column_index_from_string(cell.column)-1]] = cell.value.encode("utf-8")
                        except:
                            item['None'] = cell.value
            data.append(item)
        
        json_data = {
            "DATA_INFOS" : data,
            "DATA_NUMBER" : max_row - 1
        }

        json.dump(json_data, file, indent = 4, ensure_ascii = False)
        file.close()


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    form = request.form

    # Is the upload using Ajax, or a direct POST by the form?
    is_ajax = False
    if form.get("__ajax", None) == "true":
        is_ajax = True

    # Target folder for these uploads.
    # target = "uploadr/static/uploads/{}".format(upload_key)
    # try:
    #     os.mkdir(target)
    # except:
    #     if is_ajax:
    #         return ajax_response(False, "Couldn't create upload directory: {}".format(target))
    #     else:
    #         return "Couldn't create upload directory: {}".format(target)

    for upload in request.files.getlist("file"):
        # try:
        filename = upload.filename.rsplit("/")[0]
        filename =  filename.split('.')[0]

        wb = load_workbook(filename=upload)#, read_only=True)
        # filename = file.split('.')

        sheets = wb.get_sheet_names()
        for sheet in sheets:
            ws = wb[sheet]
            
            all_data_to_json(ws, filename, sheet)

        # except:
            # return render_template('index.html')

    return render_template('index.html')
    # try:
    #     return render_template('index.html')
    # except:
    #     return render_template('index.html')

@app.route('/search')
def search():
    files = os.listdir(DIR_JSON)
    for file in files: 
        if(".json" in file):
            print file
    return render_template('search.html')

@app.route('/download')
def download():

    files = []
    directory = os.listdir(DIR_JSON)
    
    for file in directory:
        if(".json" in file):
            files.append(file)

    return render_template('download.html', files=files)